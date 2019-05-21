import re
import PySimpleGUI as sg
import ftplib
from openpyxl import Workbook
from openpyxl.styles import Font
from ftplib import all_errors as FTP_ERRORS

totalCompanies = 0


def makeSpreadsheet():
    book = Workbook()
    company_sheet = book.active

    company_sheet['A1'].value = "First Name"
    company_sheet['A1'].font = Font(bold=True)
    company_sheet.column_dimensions['A'].width = 31.00

    company_sheet['B1'].value = "Last Name"
    company_sheet['B1'].font = Font(bold=True)
    company_sheet.column_dimensions['B'].width = 26.17

    company_sheet['C1'].value = "Company Name"
    company_sheet['C1'].font = Font(bold=True)
    company_sheet.column_dimensions['C'].width = 43.67

    company_sheet['D1'].value = "Address Line 1"
    company_sheet['D1'].font = Font(bold=True)
    company_sheet.column_dimensions['D'].width = 62.67

    company_sheet['E1'].value = "City"
    company_sheet['E1'].font = Font(bold=True)
    company_sheet.column_dimensions['E'].width = 33.67

    company_sheet['F1'].value = "State"
    company_sheet['F1'].font = Font(bold=True)
    company_sheet.column_dimensions['F'].width = 8.83

    company_sheet['G1'].value = "ZIP Code"
    company_sheet['G1'].font = Font(bold=True)
    company_sheet.column_dimensions['G'].width = 8.00

    return book, company_sheet


def exportToCSV(infoList, dest, fileName):
    global totalCompanies
    progress = 1
    book, sheet = makeSpreadsheet()
    line = 2

    for items in infoList:
        sheet.cell(line, 1).value = items["firstName"]
        sheet.cell(line, 2).value = items["lastName"]
        sheet.cell(line, 3).value = items["companyName"]
        sheet.cell(line, 4).value = items["address"]
        sheet.cell(line, 5).value = items["city"]
        sheet.cell(line, 6).value = items["state"]
        sheet.cell(line, 7).value = items["zip"]
        progress += 1
        line += 1
        bar = sg.OneLineProgressMeter('Writing to Excel', progress, totalCompanies, 'key')
        if not bar and sg.OneLineProgressMeter.exit_reasons['key'] is 'cancelled':
            totalCompanies = 0
            book.remove(sheet)
            return False

    book.save(dest + "/" + "extractedInfo_" + re.sub("c.txt","",fileName)+ ".xlsx")

    return True


def parseLine(line):
    useRA_Name = True
    infoDict = {}
    cor_number = line[0:12]
    cor_name = line[12:204]
    cor_status = line[204]
    cor_filing_type = line[205:220]
    cor_princ_add_1 = line[220:262]
    cor_princ_add_2 = line[262:304]
    cor_princ_city = line[304:332]
    cor_princ_state = line[332:334]
    cor_princ_zip = line[334:344]
    cor_princ_country = line[344:346]
    cor_mail_add_1 = line[346:388]
    cor_mail_add_2 = line[388:430]
    cor_mail_city = line[430:458]
    cor_mail_state = line[458:460]
    cor_mail_zip = line[460:470]
    cor_mail_country = line[470:472]
    cor_file_date = line[472:480]
    cor_fei_number = line[480:494]
    more_than = line[494]
    last_trx = line[495:503]
    state_country = line[503:505]
    report_year1 = line[505:509]
    house_flag1 = line[509]
    report_date1 = line[510:518]
    report_year2 = line[518:522]
    house_flag_2 = line[522]
    report_date3 = line[523:531]
    ra_name = line[531:573]
    ra_name_type = line[573]
    ra_add_1 = line[574:616]
    ra_city = line[616:644]
    ra_state = line[644:646]
    ra_zip5 = line[646:651]
    ra_zip4 = line[651:655]

    if len(line) > 650:
        princ_title = line[655:659]
        princ_name_type = line[659]
        princ_name = line[673:715]

        useRA_Name = False


    if useRA_Name == True:
        if re.search("\\s{2,}", ra_name.strip()):
            splitName = re.split('\\s{2,}', ra_name.strip())
            infoDict["lastName"] = splitName[0].strip()
            infoDict["firstName"] = splitName[1].strip()
        else:
            infoDict["firstName"] = ra_name.strip()
            infoDict["lastName"] = ""

        print(ra_name.strip() + "   RA")
    else:
        tempName = princ_name
        if re.search("\\s{2,}", tempName.strip()):
            splitName = re.split('\\s{2,}', tempName.strip())
            infoDict["lastName"] = splitName[0].strip()
            infoDict["firstName"] = splitName[1].strip()
        else:
            infoDict["firstName"] = tempName.strip()
            infoDict["lastName"] = ""

    infoDict["companyName"] = cor_name.strip()
    infoDict["address"] = cor_mail_add_1.strip()
    infoDict["city"] = cor_mail_city.strip()
    infoDict["state"] = cor_mail_state.strip()
    infoDict["zip"] = cor_mail_zip.strip()

    if infoDict["firstName"] == '' and infoDict['lastName'] == '':
        if re.search("\\s{2,}", ra_name.strip()):
            splitName = re.split('\\s{2,}', ra_name.strip())
            infoDict["lastName"] = splitName[0].strip()
            infoDict["firstName"] = splitName[1].strip()
        else:
            infoDict["firstName"] = ra_name.strip()
            infoDict["lastName"] = ""


    return infoDict


def readTextFile(newDest, key):
    index = 0
    global totalCompanies
    progress = 0
    infoList = []
    file = open(newDest, "r")

    # count companies for progress
    for i in file:
        if key == '':
            totalCompanies += 1
        elif key.upper() in i:
            totalCompanies += 1


    file.seek(0)

    if totalCompanies == 0 and key != '':
        sg.Popup("Keyword not found :(")
        return None


    for line in file:
        if key == '':
            if len(line) < 500:
                line += next(file)
            infoList.append(parseLine(line))
            progress += 1
            bar = sg.OneLineProgressMeter('Parsing Text File', progress, totalCompanies, 'key')
            if not bar and sg.OneLineProgressMeter.exit_reasons['key'] is 'cancelled':
                totalCompanies = 0
                file.close()
                return None
            index +=1
        else:
            if key.upper() in line:
                if len(line) < 500:
                    line += next(file)
                infoList.append(parseLine(line))
                progress += 1
                bar = sg.OneLineProgressMeter('Parsing Text File', progress, totalCompanies, 'key')
                if not bar and sg.OneLineProgressMeter.exit_reasons['key'] is 'cancelled':
                    totalCompanies = 0
                    file.close()
                    return None
                index +=1


    file.close()
    return infoList


def downloadTextFile(date, dest):

    path = "/public/doc/cor/"
    fileName = str(date).replace("-", "")
    fileName = fileName[0:8] + "c.txt"

    with ftplib.FTP("ftp.dos.state.fl.us") as ftp:
        ftp.login()
        ftp.cwd(path)
        newDest = dest + "/" + fileName
        with open(dest + "/" + fileName, 'wb') as fp:
            try:
                ftp.retrbinary('RETR ' + fileName, fp.write)
            except FTP_ERRORS:
                sg.Popup("Error retrieving file :(\n")

    ftp.close()

    return newDest, fileName


def runUI():
    global totalCompanies
    layout = [[sg.Text('Fetch Leads')],
              [sg.Text('Destination for Files:', size=(16, 1)), sg.InputText(key='folder'), sg.FolderBrowse(target='folder')],
              [sg.Text('Date of Desired File:', size=(16, 1)), sg.InputText('Choose the date ---->', size=(30, 1), key='date'),
               sg.CalendarButton("Date", target='date',size=(15, 1))],
              [sg.Text('Keyword:', size=(16,1)), sg.InputText('',size=(30, 1), key = 'keyword')],
              [sg.Submit()]]

    window = sg.Window('MULE 002', layout)

    while True:
        event, values = window.Read()
        if event is None:
            break
        dest = values['folder']
        date = values['date']
        key = values['keyword']

        if date == 'Choose the date ---->' or dest == '':
            sg.Popup("Please fill necessary criteria")
            continue

        newDest, fileName = downloadTextFile(date, dest)
        infoList = readTextFile(newDest, key)
        if len(infoList) != 0:
            finish = exportToCSV(infoList, dest, fileName)
            if finish is True:
                totalCompanies = 0
                sg.Popup("Done!")
                print(infoList)



def main():

    runUI()
    # dest, date = runUI()
    # newDest = downloadTextFile(date, dest)
    # infoList = readTextFile(newDest)
    # exportToCSV(infoList, dest)
    # sg.Popup("Done!")


if __name__ == "__main__":
    main()
